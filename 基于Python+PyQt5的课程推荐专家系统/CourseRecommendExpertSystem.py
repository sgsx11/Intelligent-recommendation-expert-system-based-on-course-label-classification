# -*- coding: utf-8 -*-
from itertools import combinations

import xlrd
import openpyxl


class CourseRecommendExpertSystem:

    def __init__(self,process):
        self.process = process
        #初始化规则库
        self.rule_base = []

    # 读取excel表,将前提和结论存入三维列表
    def read(self):
        #清空rule_base
        self.rule_base.clear()
        #读取规则库
        data = xlrd.open_workbook("rule_base.xlsx")
        table = data.sheets()[0]
        row = table.nrows
        col = table.ncols
        for i in range(1, row):
            rule = []
            conclusion = []
            premise = []
            for j in range(col):
                title = table.cell_value(0, j)
                if title == '结论':
                    conclusion.append(table.cell_value(i, j))
                elif title == '前提':
                    if table.cell_value(i, j) != '':
                        premise.append(table.cell_value(i, j))
            rule.append(conclusion)
            rule.append(premise)
            self.rule_base.append(rule)
        # print(self.rule_base)
    #将用户输入的规则写入规则库
    def write(self,rules):
        #print(rules)
        #读取现有规则库的行数
        wb = openpyxl.load_workbook(r'rule_base.xlsx')
        ws = wb['Sheet1']
        row = ws.max_row
        #print(ws.max_row)
        # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得
        for i in range(1, len(rules) + 1):
            rule = rules[i - 1]
            # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
            ws.cell(row=row+1, column=i).value = rule
        # 保存操作
        wb.save(r'rule_base.xlsx')
        #print("规则添加成功！")




    #求解不可重复组合数
    def combination(self,num_list):
        '''''
        生成组合,不限元素个数
        列表中元素不允许重复出现
        组合数计算为：2^n，其中n为num_list列表中元素个数
        '''
        res_list = []
        for i in range(len(num_list) + 1):
            res_list += list(combinations(num_list, i))
        del(res_list[0])#去掉空集
        #print(res_list)
        return res_list

    #查找结论
    def find_conclusion(self,facts):
        #初始化返回列表
        result = []
        #print("已知事实：{}".format(facts))
        #遍历规则库中的前提列表
        for rule in self.rule_base:
            #print("已知前提：{}".format(rule[1]))
            #判断前提列表中的前提是否都存在于facts中，则返回结论
            t = 1
            for premise in rule[1]:
                if premise not in facts:
                    t = 0
                    break
            if t:
                result.append(rule[0])#将所有结论封装到列表中返回
        #print("结论:{}".format(result))
        return result


    def inference(self,fact_list):
        self.process.append("\n开始推理......\n")
        #获取规则库
        self.read()
        #推理：（1）遍历事实列表，根据一个或多个事实(调用itertools库中的combinations方法进行组合)在规则库中查找结论
        # （2）如果找到，判断该结论是否已存在结论列表中或者事实列表中，如果为否，则将结论加入事实列表，并存入结论列表，跳转（4）
        # （3）如果没找到或者该结论已存在，继续遍历事实列表，回到（2）
        # （4）重新遍历事实列表，回到（1）
        # （5）事实列表遍历完毕，如果结论列表不为空，则取结论列表中最新的结论输出为结果
        #初始化结论列表
        conclusion_list = []
        self.recu(fact_list,conclusion_list)
        #判断推理成功还是失败
        if len(conclusion_list) != 0:
            self.process.append("\n推理成功！")
            self.process.append("结论：{}".format(conclusion_list[-1]))
            self.process.moveCursor(self.process.textCursor().End)
            #返回课程名称
            return conclusion_list[-1]
        else:
            self.process.append("\n推理失败！")
            self.process.moveCursor(self.process.textCursor().End)
            #返回False
            return False


    #递归执行推理过程
    def recu(self,fact_list,conclusion_list):
        #将事实列表中各元素进行组合
        combined_fact_list = self.combination(fact_list)
        #遍历组合过的事实列表
        for facts in combined_fact_list:
            #print(facts)
            #和规则库进行匹配，如果前提列表中的前提都存在于facts中，则查找结论成功，否则失败
            result = self.find_conclusion(facts)
            if len(result) != 0:
                #判断该结论是否已存在结论列表
                for conclusion in result:
                    if conclusion[0] not in conclusion_list and conclusion[0] not in fact_list:
                        #print("查找结论成功!")
                        self.process.append("{} --> 结论：{}".format(facts, conclusion[0]))
                        #print("该结论已存储！")
                        #将结论加入事实列表，并存入结论列表
                        fact_list.append(conclusion[0])
                        conclusion_list.append(conclusion[0])
                        #重新组合事实列表并遍历
                        self.recu(fact_list, conclusion_list)
                    else:
                        #print("{}结论已存在！".format(conclusion[0]))
                        continue
            else:
                #print("查找结论失败")
                continue